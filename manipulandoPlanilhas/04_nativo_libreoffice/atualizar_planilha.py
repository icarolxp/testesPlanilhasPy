from openpyxl import Workbook, load_workbook
import os

arquivo = 'orcamento_projeto.xlsx'

# Na vida real, você receberia esse arquivo de alguém.
# Aqui, vamos criar um rápido só para ter o que editar.
if not os.path.exists(arquivo):
    wb_temp = Workbook()
    ws_temp = wb_temp.active
    ws_temp.title = "Orçamento"
    
    # Cabeçalhos
    ws_temp['A1'] = "Item"
    ws_temp['B1'] = "Custo"
    
    # Dados iniciais
    ws_temp.append(["Servidor Fedora", 500])
    ws_temp.append(["Domínio .com", 50])
    ws_temp.append(["Licenças", 200])
    
    wb_temp.save(arquivo)
    print(f" Arquivo base '{arquivo}' criado para o teste.")

print("-" * 30)

print(f" Abrindo '{arquivo}' para atualizar valores...")

# Carrega o arquivo existente
wb = load_workbook(arquivo)
ws = wb["Orçamento"]

# 1. Atualizar um valor específico (Célula B2)
valor_antigo = ws['B2'].value
ws['B2'] = 650
print(f"   -> Item 'Servidor': Atualizado de R${valor_antigo} para R$650")

# 2. Adicionar um novo item no final da lista
ws.append(["Mão de Obra (Python)", 1500])
print("   -> Novo item adicionado: Mão de Obra")

# 3. Inserir uma fórmula de SOMA automaticamente
# ws.max_row pega o número da última linha com dados
linha_total = ws.max_row + 1
ws[f'A{linha_total}'] = "TOTAL GERAL"
ws[f'B{linha_total}'] = f"=SUM(B2:B{linha_total-1})"
print(f"   -> Fórmula de Total inserida na linha {linha_total}")

# Salva as alterações
wb.save(arquivo)
print(f" Sucesso! Abra o arquivo '{arquivo}' para ver o resultado.")