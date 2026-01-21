from openpyxl import Workbook, load_workbook
import os

arquivo = 'orcamento_projeto.xlsx'

# 1. (Criando um arquivo "falso" para editar depois)
# Na vida real, esse arquivo já existiria.
if not os.path.exists(arquivo):
    wb_temp = Workbook()
    ws_temp = wb_temp.active
    ws_temp.title = "Orçamento"
    # Cabeçalhos
    ws_temp['A1'] = "Item"
    ws_temp['B1'] = "Custo"
    # Dados originais
    ws_temp.append(["Servidor Linux", 500])
    ws_temp.append(["Domínio", 50])
    ws_temp.append(["Licenças", 200])
    wb_temp.save(arquivo)
    print(f"Arquivo base '{arquivo}' criado.")

print("-" * 30)

# 2. Editar o arquivo existente
print(f"Abrindo '{arquivo}' para edição...")

# Carregar a planilha (data_only=False para manter fórmulas se tiver)
wb = load_workbook(arquivo)
ws = wb["Orçamento"]

# Vamos atualizar o preço do "Servidor Linux" (que está na célula B2)
# Vamos supor que o preço aumentou
valor_antigo = ws['B2'].value
novo_valor = 650

ws['B2'] = novo_valor
print(f"   -> Atualizado: Servidor Linux de R${valor_antigo} para R${novo_valor}")

# Adicionar um novo item no final
ws.append(["Mão de Obra (Python)", 1500])
print("   -> Adicionado: Mão de Obra")

# Adicionar uma fórmula de SOMA no final
# Descobrir qual é a próxima linha vazia
linha_total = ws.max_row + 1
ws[f'A{linha_total}'] = "TOTAL"
ws[f'B{linha_total}'] = f"=SUM(B2:B{linha_total-1})"
print("   -> Adicionada fórmula de TOTAL")

# Salvar com o mesmo nome (sobrescrever) ou novo nome
wb.save(arquivo)
print(f"Edição concluída e salva em '{arquivo}'!")