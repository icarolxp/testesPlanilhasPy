from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# 1. Carregar o arquivo existente
nome_arquivo = 'lista_produtos.xlsx'
try:
    wb = load_workbook(nome_arquivo)
    ws = wb.active # Pega a aba ativa
except FileNotFoundError:
    print("Erro: Rode o gerador primeiro.")
    exit()

print("Editando arquivo...")

# 2. Modificando dados: Aumento de 10% nos preços

# Coluna 3 é o 'Preço_Atual' (C)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
    celula_preco = row[0] # Pega a célula da tupla
    
    # Aplica o aumento
    valor_antigo = celula_preco.value
    novo_valor = valor_antigo * 1.10
    
    celula_preco.value = novo_valor # Atualiza o valor na célula

# 3. Modificando Estilo: Destacar o Cabeçalho
# Vai pintar a primeira linha de Azul Escuro com letra Branca
azul_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
fonte_branca = Font(color="FFFFFF", bold=True)

for celula in ws[1]: # ws[1] pega toda a primeira linha
    celula.fill = azul_fill
    celula.font = fonte_branca

# 4. Adicionar uma nova coluna de "Status"
ws['E1'] = "Status Estoque" # Cabeçalho novo
ws['E1'].font = fonte_branca
ws['E1'].fill = azul_fill

# Preencher status baseado no estoque (Coluna D é a 4ª)
for linha in range(2, ws.max_row + 1):
    estoque = ws.cell(row=linha, column=4).value
    celula_status = ws.cell(row=linha, column=5)
    
    if estoque < 10:
        celula_status.value = "BAIXO"
        celula_status.font = Font(color="FF0000", bold=True) # Vermelho
    else:
        celula_status.value = "OK"

# Salvando como um novo arquivo (boa prática para não perder o original)
wb.save('lista_produtos_atualizada.xlsx')
print("Sucesso! Criado 'lista_produtos_atualizada.xlsx' com preços novos.")